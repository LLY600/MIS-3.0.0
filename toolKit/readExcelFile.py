'''
实现解析Excel文件功能
'''
#encoding = utf-8
#运行程序时不能打开Excel表

import openpyxl
import time
from FilePath.readFilePath import *
from PageObject.tableDataPage import *


class ReadExcelFile(object):

	def __init__(self, excelPath):
		self.excelPath = excelPath
		try:
			self.workbook = openpyxl.load_workbook(excelPath)  # 加载Excel
		except Exception:
			raise Exception
		self.sheet = self.workbook.active  # 获取第一个sheet

	def getSheetByIndex(self, sheetIndex):  # 设置当前要操作的sheet对象，使用index来获取
		sheet_name = self.workbook.sheetnames[sheetIndex]
		self.sheet = self.workbook[sheet_name]
		return self.sheet

	def getDefaultSheet(self):  # 获取默认的sheet名字
		return self.sheet.title

	def getSheetByName(self, sheetName):  # 设置当前要操作的sheet对象，使用sheet名称来获取
		sheet = self.workbook[sheetName]
		self.sheet = sheet
		return self.sheet

	def getMaxRowNumber(self):  # 获取最大行数
		return self.sheet.max_row

	def getMaxColNumber(self):  # 获取最大列数
		return self.sheet.max_column

	def getMinRowNumber(self):  # 获取最小行数
		return self.sheet.min_row

	def getMinColNumber(self):  # 获取最小列数
		return self.sheet.min_column

	def getAllRows(self):  # 获取默认sheet所有行对象
		return list(self.sheet.iter_rows())

	def getAllCols(self):  # 获取默认sheet所有列对象
		return list(self.sheet.iter_cols())

	def getSingleRow(self, rowNo):  # 获取默认sheet中的某一行，第一行从0开始
		return self.getAllRows()[rowNo]

	def getSingleCol(self, colNo):  # 获取默认sheet中的某一列，第一列从0开始
		return self.getAllCols()[colNo]

	def getCell(self, rowNo, colsNo):  # 获取指定单元格
		return self.sheet.cell(row=rowNo, column=colsNo)

	def getCellOfValue(self, rowNo, colsNo):  # 获取指定单元格的值
		return self.sheet.cell(row=rowNo, column=colsNo).value

	def writeValueToCell(self, content, rowNo, colsNo):  # 向指定单元格写入内容
		self.sheet.cell(row=rowNo, column=colsNo).value = content
		self.workbook.save(self.excelPath)
		return self.sheet.cell(row=rowNo, column=colsNo).value

	def writeCurrentTimeToCell(self, rowNo, colsNo):  # 向指定单元格写入时间
		now = int(time.time())
		timeArray = time.localtime(now)
		currentTime = time.strftime('%Y-%m-%d %H:%M:%S', timeArray)
		self.sheet.cell(row=rowNo, column=colsNo).value = currentTime
		self.workbook.save(self.excelPath)
		return self.sheet.cell(row=rowNo, column=colsNo).value

	def saveExcelFile(self):
		self.workbook.save(self.excelPath)

	# 获取用例起始标志位
	def get_start_sign(self, ref, startSign):
		data = ref.getAllRows()
		for row_num, row in enumerate(data):
			row_num = row_num + 1
			if row[0].value == startSign:
				return row_num

	# 获取用例结束标志位，此处以起始标志位为基准，所以参数仍是起始标志；以单元格值为None作判断
	def get_end_sign(self, ref, startSign):
		data = ref.getAllRows()
		for index, row in enumerate(data):
			index += 1
			if row[0].value == startSign:
				for i in range(index,  ref.getMaxRowNumber()+2):
					while ref.getCell(i, 1).value == None:
						return i

	# 循环列表数据次数，即用例条数
	def cycle_index_list(self, ref, startSign):
		start = ref.get_start_sign(ref, startSign)
		end = ref.get_end_sign(ref, startSign)
		cycle_index = end - start - 2
		list = []
		for i in range(cycle_index):
			list.append(ref.getSingleRow(start + i + 1))
		return list



if __name__ == '__main__':
	ref = ReadExcelFile(excelDataPath)
	ref.getSheetByName(u'基础数据')
	# print('获取默认的sheet：', ref.getDefaultSheet())
	# print('设置sheet索引为1：', ref.getSheetByIndex(1))
	# print('获取默认的sheet：', ref.getDefaultSheet())
	# print('设置sheet索引为0：', ref.getSheetByIndex(0))
	# print('获取默认的sheet：', ref.getDefaultSheet())
	print('最大行数：', ref.getMaxRowNumber())
	# print('最大列数：', ref.getMaxColNumber())
	# print('最小行数：', ref.getMinRowNumber())
	# print('最小列数：', ref.getMinColNumber())
	# print('所有行内容：', ref.getAllRows())  # 从左至右读取
	# print('所有列内容：', ref.getAllCols())  # 从上至下读取
	# print('获取某一列内容：', ref.getSingleCol(2))
	# print('获取某一行内容：', ref.getSingleRow(10))
	# print('获取行列为（1,1）的单元格：', ref.getCell(1, 1))
	# print('获取单元格（1,1）的值：', ref.getCellOfValue(1, 1))  # 下标从1开始
	# print('单元格（10,10）写入“thinker”：', ref.writeValueToCell('thinker', 10, 10))
	# print('单元格（13,13）写入当前时间：', ref.writeCurrentTimeToCell(13, 13))
	# ref.getSheetByName('登录账号')
	# # print(ref.getAllRows())
	# print('除去标题列表数据：', ref.getAllRows()[1:])
	# rows_login = ref.getAllRows()[1:]
	# list = [ref.getSingleRow(9), ref.getSingleRow(10)]
	# for index01, row01 in enumerate(list):
	# 	print(index01)
	# 	print(row01)
	# 	# print(index01, row01)
	# 	print(row01[1].value, row01[6].value)
	# ref.getExcelSign(u'添加物业')
	ref.getEndSign(ref.getSheetByName(u'基础数据'), u'添加物业')


